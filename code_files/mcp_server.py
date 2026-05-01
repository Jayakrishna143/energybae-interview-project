"""
mcp_server.py — Execution / MCP Server  (FastAPI on port 8001)
----------------------------------------------------------------
Responsible ONLY for the safe Excel write operation.

Excel structure (from "Copy of Pranay HOME E-Bill Analysis.xlsx"):
  Row 1  : Consumer Name   → D1 (consumer 1) | H1 (consumer 2)
  Row 2  : Consumer No     → D2              | H2
  Row 3  : Fixed Charges   → D3              | H3
  Row 8  : Headers         → C=Month, D=Units, E=Bill Amt, F=Unit Cost (formula)
                           → G=Month, H=Units, I=Bill Amt, J=Unit Cost (formula)
  Row 9+ : Monthly data rows

  ✅  We write to: C, D, E (consumer 1) or G, H, I (consumer 2) + Sr.No in B
  ❌  We NEVER touch: F (=(E-D3)/D) or J (=(I-H3)/H) — formulas are preserved
"""

import os
from datetime import datetime
from pathlib import Path

import openpyxl
from fastapi import FastAPI
from pydantic import BaseModel

app = FastAPI(title="MCP Excel Execution Server")

# The Excel file must be in the same directory as this script (or CWD).
EXCEL_FILE = "Copy of Pranay HOME E-Bill Analysis.xlsx"

# Column indices (1-based, as openpyxl uses)
COL = {
    "sr_no":           2,   # B — shared by both consumers
    # Consumer 1 (Madhusham — higher consumer number)
    "c1_month":        3,   # C
    "c1_units":        4,   # D
    "c1_bill":         5,   # E
    "c1_unit_cost":    6,   # F  ← FORMULA — do not overwrite, just set formula for new rows
    # Consumer 2 (Ranjana — lower consumer number)
    "c2_month":        7,   # G
    "c2_units":        8,   # H
    "c2_bill":         9,   # I
    "c2_unit_cost":   10,   # J  ← FORMULA
}

HEADER_ROW  = 8   # row with column labels
DATA_START  = 9   # first actual data row


class BillData(BaseModel):
    customer_name:      str
    consumer_number:    str
    sanctioned_load_kw: float
    tariff_category:    str
    units_consumed:     int
    bill_amount:        float
    bill_month:         str    # "YYYY-MM"


# ── Helper: find which consumer this bill belongs to ──────────────────────────
def identify_consumer(ws, consumer_number: str) -> int:
    """Returns 1 or 2 based on matching consumer number in D2 / H2."""
    # D2 and H2 are stored as floats (e.g. 439320095567.0) — normalise to string
    c1 = str(int(ws["D2"].value)) if ws["D2"].value else ""
    c2 = str(int(ws["H2"].value)) if ws["H2"].value else ""
    clean = consumer_number.strip()
    if clean == c1:
        return 1
    if clean == c2:
        return 2
    raise ValueError(
        f"Consumer number {clean} not found in Excel.\n"
        f"  Expected {c1} (consumer 1) or {c2} (consumer 2)."
    )


# ── Helper: find the next empty row for a given unit-column ──────────────────
def find_next_row(ws, unit_col: int) -> int:
    """Scan from DATA_START downward; return first row where unit_col is empty."""
    for row in range(DATA_START, DATA_START + 200):
        if ws.cell(row=row, column=unit_col).value is None:
            return row
    raise RuntimeError("Could not find an empty row within 200 rows of data.")


# ── Helper: get the last Sr.No value ─────────────────────────────────────────
def get_last_sr_no(ws) -> int:
    last = 1
    for row in range(DATA_START, DATA_START + 200):
        v = ws.cell(row=row, column=COL["sr_no"]).value
        if v is not None:
            try:
                last = int(v)
            except (TypeError, ValueError):
                pass
    return last


# ═══════════════════════════════════════════════════════════════════════════════
#  TOOL: update_solar_excel
#  This is the "MCP tool" — the only function that touches the workbook.
# ═══════════════════════════════════════════════════════════════════════════════
@app.post("/update-excel")
def update_solar_excel(data: BillData):
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        return {"success": False, "error": f"Excel file not found: {EXCEL_FILE}"}

    try:
        # ── Load workbook WITHOUT data_only so formulas are preserved ────────
        # data_only=False (default) keeps formula strings intact.
        # If we had used data_only=True and then saved, formulas would be lost.
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # ── Identify consumer ─────────────────────────────────────────────────
        consumer = identify_consumer(ws, data.consumer_number)

        if consumer == 1:
            col_month     = COL["c1_month"]
            col_units     = COL["c1_units"]
            col_bill      = COL["c1_bill"]
            col_unit_cost = COL["c1_unit_cost"]
            fixed_charge_cell = "$D$3"   # used in unit-cost formula
        else:
            col_month     = COL["c2_month"]
            col_units     = COL["c2_units"]
            col_bill      = COL["c2_bill"]
            col_unit_cost = COL["c2_unit_cost"]
            fixed_charge_cell = "$H$3"

        # ── Find the target row ───────────────────────────────────────────────
        target_row = find_next_row(ws, col_units)

        # ── Parse bill_month into a datetime (Excel stores dates as datetimes) ─
        month_dt = datetime.strptime(data.bill_month, "%Y-%m")

        # ── Write Sr.No (only if the B column for this row is empty) ─────────
        if ws.cell(row=target_row, column=COL["sr_no"]).value is None:
            ws.cell(row=target_row, column=COL["sr_no"]).value = get_last_sr_no(ws) + 1

        # ── Write Month ───────────────────────────────────────────────────────
        ws.cell(row=target_row, column=col_month).value = month_dt

        # ── Write Units ───────────────────────────────────────────────────────
        ws.cell(row=target_row, column=col_units).value = data.units_consumed

        # ── Write Bill Amount ─────────────────────────────────────────────────
        ws.cell(row=target_row, column=col_bill).value = data.bill_amount

        # ── Write Unit Cost formula (mirrors the pattern in row 20) ──────────
        # Formula pattern: =(E20-$D$3)/D20  →  =(E{row}-{fixed}-{units_col}{row})
        unit_letter  = openpyxl.utils.get_column_letter(col_units)
        bill_letter  = openpyxl.utils.get_column_letter(col_bill)
        formula = f"=({bill_letter}{target_row}-{fixed_charge_cell})/{unit_letter}{target_row}"
        ws.cell(row=target_row, column=col_unit_cost).value = formula

        # ── Save (openpyxl preserves all other formulas we didn't touch) ──────
        wb.save(excel_path)

        return {
            "success":  True,
            "message":  f"Row {target_row} updated for {data.customer_name} (consumer {consumer})",
            "row":      target_row,
            "consumer": consumer,
        }

    except Exception as exc:
        return {"success": False, "error": str(exc)}


@app.get("/health")
def health():
    return {"status": "mcp-server ok"}
